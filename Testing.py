import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side  # Add these imports
from tkinter import ttk



def load_menu():
    # Load menu from JSON file or any other source
    # For simplicity, we'll use a sample menu here
    return {
        "starters": [
            {"item": "Tzatziki", "price": 3.70},
            {"item": "Tirokafteri", "price": 3.70},
            {"item": "Spicy Cheese Sauce", "price": 3.70},
            {"item": "Humus", "price": 3.70},
            {"item": "Oven Fetta", "price": 6.80},
            {"item": "Fetta", "price": 3.80},
            {"item": "Fries", "price": 3.80},
            {"item": "Fries with Cheese", "price": 4.50},
            {"item": "Fries with Fetta & Oregano", "price": 5.50},
            {"item": "Olives", "price": 3.80}
        ],
        "salads": [
            {"item": "Greek Salad", "price": 8.20},
            {"item": "Dakos", "price": 6.20},
            {"item": "Chicken Salad", "price": 8.40}
        ],
        "pitta_wraps": [
            {"item": "Pitta Yeeros Pork", "price": 7.50},
            {"item": "Pitta Yeeros Chicken", "price": 7.50},
            {"item": "Pitta Yeeros Mix Pork & Chicken", "price": 7.80},
            {"item": "Pitta Skewers Pork", "price": 7.35}
        ],
        "portions": [
            {"item": "Yeeros Pork", "price": 14.95},
            {"item": "Yeeros Chicken", "price": 14.70},
            {"item": "Yeeros Mix Pork & Chicken", "price": 14.95},
            {"item": "Skewers Pork", "price": 12.50}
        ],
        "yeeros_burgers": [
            {"item": "Chicken Yeeros Burger", "price": 9.90},
            {"item": "Pork Yeeros Burger", "price": 9.90},
            {"item": "Mixed Yeeros Burger (Pork & Chicken)", "price": 10.40}
        ],
        "skepasti": [
            {"item": "Chicken", "price": 15.50},
            {"item": "Pork", "price": 15.80}

        ],
        "mixed_grill": [
            {"item": "Mixed Grill for 2", "price": 29.90},
            {"item": "Fully Combo Mix Grill", "price": 35.90}

        ],
        "halal": [
            {"item": "Chicken Skewers", "price": 12.50},
            {"item": "Kebab Portion", "price": 12.50}

        ],
        "on_its_own": [
            {"item": "Pork Skewer", "price": 3.30},
            {"item": "Chicken Skewer", "price": 3.30}

        ],
        "desserts": [
            {"item": "Ferrero", "price": 5.30},
            {"item": "Cheese Cake", "price": 5.30},
            {"item": "Ekmek Kataifi", "price": 5.30}
        ],
        "soft_drinks": [
            {"item": "Water", "price": 1.80},
            {"item": "Coka Cola (Zero, Diet)", "price": 1.80}
        ],
        "kids_menu": [
            {"item": "Mini Pitta Yeeros Chicken", "price": 5.10},
            {"item": "Mini Pitta Yeeros Pork", "price": 5.10}
        ]
    }


class RestaurantApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Mr. Souvlaki Restaurant")
        self.root.geometry("1200x800")
        self.root.bind('<Return>', lambda event: self.process_order())


        self.menu = load_menu()
        self.orders = []  # List to store multiple orders
        self.current_order = {"items": [], "total_amount": 0}  # Current order

        self.order_number = 1  # Counter for order number

        self.frames = []
        self.create_frames()

        # GUI Components
        self.label_menu = tk.Label(root, text="Add item:")
        self.label_menu.grid(row=3, column=2)

        # Use Entry for direct input of item numbers
        self.entry_menu = tk.Entry(root)
        self.entry_menu.grid(row=4, column=2)

        self.button_order = tk.Button(root, text="Order", command=self.process_order)
        self.button_order.grid(row=5, column=2)

        self.label_summary = tk.Label(root, text="Order Summary:")
        self.label_summary.grid(row=3, column=0)

        self.listbox_summary = tk.Listbox(root, height=10, width=50)
        self.listbox_summary.grid(row=4, column=0)

        self.button_display_summary = tk.Button(root, text="Display Summary", command=self.display_summary)
        self.button_display_summary.grid(row=6, column=0)

        self.label_current_order = tk.Label(root, text="Current Order:")
        self.label_current_order.grid(row=3, column=1)

        self.listbox_current_order = tk.Listbox(root, height=10, width=50)
        self.listbox_current_order.grid(row=4, column=1)

        self.button_remove_item = tk.Button(root, text="Remove Item", command=self.remove_item)
        self.button_remove_item.grid(row=6, column=2)

        self.button_generate_excel = tk.Button(root, text="Generate Excel", command=self.generate_excel)
        self.button_generate_excel.grid(row=6, column=1)

        self.button_new_order = tk.Button(root, text="New Order", command=self.new_order)
        self.button_new_order.grid(row=7, column=2)


  



        self.workbook = Workbook()
        self.excel_sheet = self.workbook.active
        self.excel_sheet.title = "Order Details"

        # Set column headers and format
        self.excel_sheet["A1"] = "Item"
        self.excel_sheet["B1"] = "Price"
        self.excel_sheet["A1"].font = Font(bold=True)
        self.excel_sheet["B1"].font = Font(bold=True)
        self.excel_sheet.column_dimensions['A'].width = 30

        self.load_menu_items()
    

    def create_frames(self):
        num_columns = 4
        num_rows_special_category = 4

        frame_index = 0
        current_row = 0

        cumulative_index = 1

        for category, items in self.menu.items():
            frame = tk.Frame(self.root)
            frame.grid(row=current_row, column=frame_index)
            self.frames.append(frame)

            label = tk.Label(frame, text=f"{category.capitalize()}:")
            label.pack(side=tk.TOP)

            listbox = tk.Listbox(frame, height=12, width=40)  # Adjust the width as needed
            listbox.pack(side=tk.TOP)

            num_items = len(items)
            items_per_column = (num_items + 3) // 4  # Ceiling division to ensure enough rows

            for col in range(4):
                start_idx = col * items_per_column
                end_idx = min((col + 1) * items_per_column, num_items)

                for idx, item in enumerate(items[start_idx:end_idx], start=start_idx + 1):
                    listbox.insert(tk.END, f"({cumulative_index}) {item['item']} - £{item['price']:.2f}")
                    cumulative_index += 1

            frame_index += 1

            if frame_index % num_columns == 0:
                current_row += 1
                frame_index = 0

            if current_row == num_rows_special_category and category == 'special_category':
                tk.Label(self.root, text="").grid(row=current_row, column=frame_index)
                current_row = 0

    def load_menu_items(self):
        for category, items in self.menu.items():
            frame = getattr(self, f"frame_{category.lower()}", None)
            if frame:
                label = tk.Label(frame, text=f"{category.capitalize()}:")
                label.pack()
                listbox = tk.Listbox(frame, height=10, width=50)
                listbox.pack()
                cumulative_index = 1
                for idx, item in enumerate(items, 1):
                    listbox.insert(tk.END, f"({cumulative_index}) {item['item']} - £{item['price']:.2f}")
                    cumulative_index += 1

    def process_order(self):
        item_number = self.entry_menu.get()
        if item_number.isdigit():
            item_number = int(item_number)
            selected_item = self.get_selected_item(item_number)
            if selected_item:
                self.current_order["items"].append(selected_item)
                self.current_order["total_amount"] += selected_item["price"]
                self.update_current_order_display()

            else:
                messagebox.showwarning("Invalid Selection", "Invalid item number. Please try again.")
        else:
            messagebox.showwarning("Invalid Input", "Please enter a valid item number.")

    def get_selected_item(self, selected_index):
        current_index = 0
        for items in self.menu.values():
            if current_index + len(items) >= selected_index + 1:
                return items[selected_index - current_index - 1]
            current_index += len(items)
        return None

    def remove_item(self):
        selected_index = self.listbox_current_order.curselection()
        if selected_index:
            selected_index = int(selected_index[0])
            removed_item = self.current_order["items"].pop(selected_index)
            self.current_order["total_amount"] -= removed_item["price"]
            self.update_current_order_display()
            messagebox.showinfo("Item Removed", f"Item '{removed_item['item']}' removed from order {self.order_number}!")
        else:
            messagebox.showwarning("No Item Selected", "Please select an item to remove.")

    def update_current_order_display(self):
        self.listbox_current_order.delete(0, tk.END)
        for item in self.current_order["items"]:
            self.listbox_current_order.insert(tk.END, f"{item['item']} - £{item['price']:.2f}")
        self.listbox_current_order.insert(tk.END, f"Total Amount: £{self.current_order['total_amount']:.2f}")

    def new_order(self):
        self.orders.append(self.current_order)
        self.current_order = {"items": [], "total_amount": 0}
        self.update_order_summary()
        self.order_number += 1
        messagebox.showinfo("New Order", f"Starting new order: Order {self.order_number}")

    def update_order_summary(self):
        self.listbox_summary.delete(0, tk.END)
        for idx, order in enumerate(self.orders, start=1):
            self.listbox_summary.insert(tk.END, f"Order {idx} - Total Amount: £{order['total_amount']:.2f}")

    def display_summary(self):
        selected_index = self.listbox_summary.curselection()
        if selected_index:
            selected_index = int(selected_index[0])
            selected_order = self.orders[selected_index]
            self.listbox_current_order.delete(0, tk.END)
            for item in selected_order["items"]:
                self.listbox_current_order.insert(tk.END, f"{item['item']} - £{item['price']:.2f}")
            self.listbox_current_order.insert(tk.END, f"Total Amount: £{selected_order['total_amount']:.2f}")
        else:
            messagebox.showwarning("No Order Selected", "Please select an order to display.")



    def generate_excel(self):
        if not self.orders:
            messagebox.showwarning("No Orders", "There are no orders to generate Excel file.")
            return

        # Create a new workbook
        workbook = Workbook()
        sheet = workbook.active

        # Define font and alignment styles
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        item_alignment = Alignment(wrap_text=True)
        
        # Define border styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Set header row style
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add headers to the sheet
        sheet.append(["Order Number", "Item", "Price"])

        # Iterate over each order
        for order_index, order in enumerate(self.orders, start=1):
            # Iterate over each item in the order
            for item in order["items"]:
                sheet.append([order_index, item["item"], item["price"]])

        # Apply styles to item rows
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = item_alignment
                cell.border = thin_border

        # Add a total amount row for each order
        for order_index, order in enumerate(self.orders, start=1):
            sheet.append(["Total Amount", "", order["total_amount"]])
            last_row = sheet.max_row

            # Merge cells for the total amount row
            sheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=2)

            # Apply styles to the total amount row
            for cell in sheet[last_row]:
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
                cell.border = thin_border

        # Adjust column widths
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 15

        # Save the workbook to a file
        filename = "orders_excel.xlsx"
        workbook.save(filename)
        messagebox.showinfo("Excel Generated", f"Excel file '{filename}' generated successfully!")


if __name__ == "__main__":
    root = tk.Tk()
    app = RestaurantApp(root)
    root.mainloop()
